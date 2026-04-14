"""Import client-jargon → product_id mappings into knowledge_product_aliases.

Reads an xlsx file with sheets containing columns
  'Запрос клиента / Заказ' | 'Позиция в 1с'
and links each (client query, 1C position) pair to a row in products.

Matching strategy:
  1. exact (normalized name) — confidence = 1.0
  2. fuzzy via BGE-M3 embeddings — confidence = 1 - distance, kept if >= FUZZY_THRESHOLD

Manual entries (match_type='manual') in the destination table are NEVER
overwritten by this script — it uses INSERT OR IGNORE on (alias_norm, product_id).

Usage:
    python scripts/import_knowledge_aliases.py \
        --input ~/Downloads/Комплектуха для ИИ.xlsx \
        --source-tag xlsx:Комплектуха \
        [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import logging
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("import_knowledge_aliases")

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "chatapp_data_prod.db"
OUT_DIR = ROOT / "scripts" / "out"

FUZZY_THRESHOLD = 0.85
SKIP_POS_PREFIXES = ("это обозначение", "это ")  # meta rows in sheet 19.11


def _norm(s: str) -> str:
    if not s:
        return ""
    s = s.lower().replace("ё", "е").strip()
    s = re.sub(r"\s*№\s*", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _ensure_table(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS knowledge_product_aliases (
          id               INTEGER PRIMARY KEY AUTOINCREMENT,
          alias            TEXT NOT NULL,
          alias_norm       TEXT NOT NULL,
          product_id       INTEGER NOT NULL REFERENCES products(id),
          match_type       TEXT NOT NULL,
          match_confidence REAL,
          source           TEXT NOT NULL,
          source_row_ref   TEXT,
          note             TEXT,
          created_at       TEXT NOT NULL DEFAULT (datetime('now')),
          updated_at       TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS ix_kpa_alias_norm ON knowledge_product_aliases(alias_norm);
        CREATE INDEX IF NOT EXISTS ix_kpa_product_id ON knowledge_product_aliases(product_id);
        CREATE UNIQUE INDEX IF NOT EXISTS ux_kpa_alias_product
            ON knowledge_product_aliases(alias_norm, product_id);
        """
    )
    conn.commit()


def _read_pairs(xlsx_path: Path, sheet_limit: int | None = 4) -> list[tuple[str, str, str, str]]:
    """Return list of (sheet_name, row_ref, client_query, position_1c)."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    sheets = wb.sheetnames
    if sheet_limit:
        sheets = sheets[:sheet_limit]

    pairs: list[tuple[str, str, str, str]] = []
    for sn in sheets:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # header: col 0 = client query, col 2 = 1C position (verified externally)
        cur_q: str | None = None
        for row_ix, r in enumerate(rows[1:], start=2):
            q = r[0] if len(r) > 0 else None
            pos = r[2] if len(r) > 2 else None
            if q:
                cur_q = str(q).strip()
            if pos and cur_q:
                pos_s = str(pos).strip()
                if pos_s.lower().startswith(SKIP_POS_PREFIXES):
                    continue
                pairs.append((sn, f"{sn}:{row_ix}", cur_q, pos_s))
    return pairs


def _dedup(
    pairs: list[tuple[str, str, str, str]],
) -> list[tuple[str, str, str, str]]:
    seen: set[tuple[str, str]] = set()
    out: list[tuple[str, str, str, str]] = []
    for sn, ref, q, pos in pairs:
        key = (_norm(q), _norm(pos))
        if key in seen:
            continue
        seen.add(key)
        out.append((sn, ref, q, pos))
    return out


def _load_products(conn: sqlite3.Connection) -> tuple[dict[str, int], list[tuple[int, str]]]:
    rows = conn.execute("SELECT id, name FROM products").fetchall()
    exact_idx: dict[str, int] = {}
    for r in rows:
        exact_idx[_norm(r[1])] = r[0]
    return exact_idx, [(r[0], r[1]) for r in rows]


def _fuzzy_match(
    positions: list[str],
    products: list[tuple[int, str]],
    threshold: float,
) -> dict[str, tuple[int, float] | None]:
    """Embed all query positions + all product names once, return best match per position."""
    if not positions:
        return {}

    logger.info(f"Loading BGE-M3 for fuzzy match on {len(positions)} positions...")
    import warnings
    warnings.filterwarnings("ignore")
    import numpy as np
    import torch
    from sentence_transformers import SentenceTransformer

    device = "mps" if torch.backends.mps.is_available() else "cpu"
    model = SentenceTransformer("BAAI/bge-m3", device=device)

    prod_names = [n for _, n in products]
    prod_ids = [i for i, _ in products]

    prod_vecs = model.encode(prod_names, normalize_embeddings=True, show_progress_bar=False)
    q_vecs = model.encode(positions, normalize_embeddings=True, show_progress_bar=False)

    # cosine similarity = dot product (already normalized)
    sims = np.asarray(q_vecs) @ np.asarray(prod_vecs).T  # [Q, P]
    best_ix = sims.argmax(axis=1)
    best_sim = sims.max(axis=1)

    result: dict[str, tuple[int, float] | None] = {}
    for i, pos in enumerate(positions):
        conf = float(best_sim[i])
        if conf >= threshold:
            result[pos] = (int(prod_ids[best_ix[i]]), conf)
        else:
            result[pos] = None
    return result


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path, help="Path to xlsx source")
    ap.add_argument("--source-tag", required=True, help="Short source id, e.g. 'xlsx:Комплектуха'")
    ap.add_argument("--sheets", type=int, default=4, help="Number of sheets from start to read (default 4)")
    ap.add_argument("--fuzzy-threshold", type=float, default=FUZZY_THRESHOLD)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    xlsx = args.input.expanduser().resolve()
    if not xlsx.exists():
        logger.error(f"input not found: {xlsx}")
        sys.exit(1)
    if not DB_PATH.exists():
        logger.error(f"db not found: {DB_PATH}")
        sys.exit(1)

    logger.info(f"reading {xlsx.name}")
    raw_pairs = _read_pairs(xlsx, sheet_limit=args.sheets)
    logger.info(f"raw pairs: {len(raw_pairs)}")
    pairs = _dedup(raw_pairs)
    logger.info(f"after dedup: {len(pairs)}")

    conn = sqlite3.connect(str(DB_PATH))
    _ensure_table(conn)
    exact_idx, products = _load_products(conn)
    logger.info(f"products in DB: {len(products)}")

    # Phase 1: exact
    exact: list[tuple[str, str, str, int]] = []
    need_fuzzy: list[tuple[str, str, str, str]] = []
    for sn, ref, q, pos in pairs:
        pid = exact_idx.get(_norm(pos))
        if pid is not None:
            exact.append((ref, q, pos, pid))
        else:
            need_fuzzy.append((sn, ref, q, pos))
    logger.info(f"exact matches: {len(exact)}")
    logger.info(f"need fuzzy:    {len(need_fuzzy)}")

    # Phase 2: fuzzy
    uniq_pos_for_fuzzy = sorted({pos for _, _, _, pos in need_fuzzy})
    fuzzy_map = _fuzzy_match(uniq_pos_for_fuzzy, products, args.fuzzy_threshold) if uniq_pos_for_fuzzy else {}

    fuzzy_hits: list[tuple[str, str, str, int, float]] = []
    unmatched: list[tuple[str, str, str, str, float | None]] = []
    for sn, ref, q, pos in need_fuzzy:
        m = fuzzy_map.get(pos)
        if m is None:
            # get best sim even below threshold for the report
            unmatched.append((sn, ref, q, pos, None))
        else:
            pid, conf = m
            fuzzy_hits.append((ref, q, pos, pid, conf))
    logger.info(f"fuzzy accepted (>= {args.fuzzy_threshold}): {len(fuzzy_hits)}")
    logger.info(f"unmatched:     {len(unmatched)}")

    # Phase 3: write
    if args.dry_run:
        logger.info("DRY RUN — no DB changes")
    else:
        cur = conn.cursor()
        ins_exact = ins_fuzzy = 0
        for ref, q, pos, pid in exact:
            try:
                cur.execute(
                    """INSERT INTO knowledge_product_aliases
                       (alias, alias_norm, product_id, match_type, match_confidence, source, source_row_ref)
                       VALUES (?, ?, ?, 'exact', 1.0, ?, ?)""",
                    (q, _norm(q), pid, args.source_tag, ref),
                )
                ins_exact += cur.rowcount
            except sqlite3.IntegrityError:
                pass
        for ref, q, pos, pid, conf in fuzzy_hits:
            try:
                cur.execute(
                    """INSERT INTO knowledge_product_aliases
                       (alias, alias_norm, product_id, match_type, match_confidence, source, source_row_ref, note)
                       VALUES (?, ?, ?, 'fuzzy', ?, ?, ?, ?)""",
                    (q, _norm(q), pid, conf, args.source_tag, ref, f"matched via BGE-M3 to: {pos}"),
                )
                ins_fuzzy += cur.rowcount
            except sqlite3.IntegrityError:
                pass
        conn.commit()
        logger.info(f"inserted exact: {ins_exact}")
        logger.info(f"inserted fuzzy: {ins_fuzzy}")

    # Report
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = OUT_DIR / "knowledge_aliases_import_report.csv"
    with report_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["source_row", "alias", "matched_product_id", "matched_product_name", "match_type", "confidence"])
        prod_name_by_id = {i: n for i, n in products}
        for ref, q, pos, pid in exact:
            w.writerow([ref, q, pid, prod_name_by_id.get(pid, ""), "exact", 1.0])
        for ref, q, pos, pid, conf in fuzzy_hits:
            w.writerow([ref, q, pid, prod_name_by_id.get(pid, ""), "fuzzy", f"{conf:.4f}"])
        for sn, ref, q, pos, _ in unmatched:
            w.writerow([ref, q, "", pos, "unmatched", ""])
    logger.info(f"\nreport: {report_path}")

    # Final stats from DB
    if not args.dry_run:
        total = conn.execute("SELECT COUNT(*) FROM knowledge_product_aliases").fetchone()[0]
        by_type = conn.execute(
            "SELECT match_type, COUNT(*) FROM knowledge_product_aliases GROUP BY match_type"
        ).fetchall()
        logger.info(f"\ntotal rows in knowledge_product_aliases: {total}")
        for t, c in by_type:
            logger.info(f"  {t}: {c}")
    conn.close()


if __name__ == "__main__":
    main()
