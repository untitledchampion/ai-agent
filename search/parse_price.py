"""Parse the Excel price list into a new `products` table in chatapp_data.db.

Adds a NEW table only — never touches existing collector tables.
"""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

PRICE_XLSX = Path("/tmp/price_extract/Прайс_Прайс-Лист 11 апреля 2026_г..xlsx")
DB_PATH = Path(__file__).resolve().parent.parent / "data" / "chatapp_data.db"

# Standard sheet column layout (0-based), header at row 8
STD_COLS = {
    "name": 0,
    "url": 5,
    "unit": 6,
    "code": 9,
    "price_dealer": 11,
    "price_small": 12,
    "price_large": 13,
}

POLOTNA_COLS = {
    "name": 0,
    "color": 2,
    "width": 3,
    "price_otrez": 5,
    "price_garpun": 7,
}

STD_SHEETS = [
    "КОМПЛЕКТУЮЩИЕ",
    "ПРОФИЛИ",
    "ИНСТРУМЕНТ",
    "РАСХОДНЫЕ МАТЕРИАЛЫ",
    "СВЕТОТЕХНИКА",
    "ГАРДИНЫ ПВХ",
]


def _clean_price(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).replace("\xa0", " ").strip()
    return s or None


def parse_standard(ws, sheet_name: str) -> list[dict]:
    out: list[dict] = []
    current_category: str | None = None
    for row in ws.iter_rows(min_row=9, values_only=True):
        if not row or all(c is None for c in row):
            continue
        name = _clean_str(row[STD_COLS["name"]])
        if not name:
            continue
        code = _clean_str(row[STD_COLS["code"]])
        if code is None:
            # Category/section header line — remember it for subsequent products
            current_category = name
            continue
        out.append({
            "sheet": sheet_name,
            "category": current_category,
            "name": name,
            "code": code,
            "unit": _clean_str(row[STD_COLS["unit"]]),
            "url": _clean_str(row[STD_COLS["url"]]),
            "price_dealer": _clean_price(row[STD_COLS["price_dealer"]]),
            "price_small": _clean_price(row[STD_COLS["price_small"]]),
            "price_large": _clean_price(row[STD_COLS["price_large"]]),
            "color": None,
            "width": None,
        })
    return out


def parse_polotna(ws) -> list[dict]:
    out: list[dict] = []
    current_category: str | None = None
    for row in ws.iter_rows(min_row=9, values_only=True):
        if not row or all(c is None for c in row):
            continue
        name = _clean_str(row[POLOTNA_COLS["name"]])
        if not name:
            continue
        price_otrez = _clean_price(row[POLOTNA_COLS["price_otrez"]])
        if price_otrez is None:
            current_category = name
            continue
        out.append({
            "sheet": "ПОЛОТНА",
            "category": current_category,
            "name": name,
            "code": None,
            "unit": "м2",
            "url": None,
            "price_dealer": price_otrez,
            "price_small": _clean_price(row[POLOTNA_COLS["price_garpun"]]),
            "price_large": None,
            "color": _clean_str(row[POLOTNA_COLS["color"]]),
            "width": _clean_str(row[POLOTNA_COLS["width"]]),
        })
    return out


def main() -> int:
    if not PRICE_XLSX.exists():
        print(f"price file not found: {PRICE_XLSX}", file=sys.stderr)
        return 1

    wb = load_workbook(PRICE_XLSX, data_only=True, read_only=True)
    products: list[dict] = []
    products.extend(parse_polotna(wb["ПОЛОТНА"]))
    for sheet in STD_SHEETS:
        products.extend(parse_standard(wb[sheet], sheet))

    print(f"parsed {len(products)} products")
    by_sheet: dict[str, int] = {}
    for p in products:
        by_sheet[p["sheet"]] = by_sheet.get(p["sheet"], 0) + 1
    for s, n in by_sheet.items():
        print(f"  {s}: {n}")

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("DROP TABLE IF EXISTS products")
        conn.execute(
            """
            CREATE TABLE products (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet         TEXT NOT NULL,
                category      TEXT,
                name          TEXT NOT NULL,
                code          TEXT,
                unit          TEXT,
                url           TEXT,
                price_dealer  REAL,
                price_small   REAL,
                price_large   REAL,
                color         TEXT,
                width         TEXT
            )
            """
        )
        conn.execute("CREATE INDEX idx_products_code ON products(code)")
        conn.execute("CREATE INDEX idx_products_sheet ON products(sheet)")
        conn.executemany(
            """
            INSERT INTO products
              (sheet, category, name, code, unit, url,
               price_dealer, price_small, price_large, color, width)
            VALUES
              (:sheet, :category, :name, :code, :unit, :url,
               :price_dealer, :price_small, :price_large, :color, :width)
            """,
            products,
        )
        conn.commit()
    finally:
        conn.close()

    print(f"wrote {len(products)} rows to {DB_PATH}:products")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
