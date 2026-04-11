#!/usr/bin/env python3
"""Build order items dictionary: aggregate raw extracted positions into Excel + CSV.

Sheets/files (NO chat_id per user request):
  - Группы:     group_key | total | variants | unique_msgs | top_members | sizes | colors | units
  - Словарь:    raw_name | total_count | unique_msgs | high/med/low | top_units
  - Все позиции: msg_id | raw_name | quantity | unit | confidence
  - Статистика:  KPI
"""
import re
import sqlite3
import csv
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DB = "/Users/klim/Desktop/ai-agent/data/chatapp_data.db"
OUT_DIR = "/Users/klim/Desktop/ai-agent/data/order_extraction"

# ============= Fuzzy clustering helpers =============

def lev(a: str, b: str) -> int:
    if a == b:
        return 0
    if len(a) < len(b):
        a, b = b, a
    if len(b) == 0:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i] + [0] * len(b)
        for j, cb in enumerate(b, 1):
            curr[j] = min(curr[j-1] + 1, prev[j] + 1, prev[j-1] + (0 if ca == cb else 1))
        prev = curr
    return prev[-1]

ENDINGS = sorted([
    "ами","ями","иях","ях","ах","ов","ев","ей","ем","ом","ых","их",
    "ай","ей","ой","ый","ий","ая","яя","ое","ее","ую","юю","ою","ею",
    "ам","ям","ою","ею","ии",
    "а","я","ы","и","у","ю","о","е","ь","й",
], key=len, reverse=True)

def stem(word: str) -> str:
    w = word
    for e in ENDINGS:
        if len(w) > len(e) + 2 and w.endswith(e):
            return w[: -len(e)]
    return w

LETTERS_RE = re.compile(r"[a-zа-яё]+", re.IGNORECASE)
DIGITS_RE = re.compile(r"\d+(?:[.,]\d+)?")
SIZE_RE = re.compile(r"\d+[./x×хЧ*]\d+", re.IGNORECASE)

UNITS_WORDS = {"шт","штук","штуки","штука","штучек","штучка","м","мм","см","кг","гр","мл","л","уп","упаковка",
               "пачка","пачку","палка","палок","палки","коробка","коробку","коробок","компл","комплект",
               "пара","пары","пар","метр","метра","метров","мп","пм"}
COLORS = {"белый","белая","белое","белых","белым","белую",
          "чёрный","черный","черная","чёрная","чёрное","черное","чёрных","черных","чёрного","черного","черную","чёрную",
          "серый","серая","серое","серых","серого",
          "красный","красная","синий","синяя","синие","синих",
          "мат","матовый","матовая","матовое","матовых","матов","мата",
          "сатин","сатиновый","глянец","глянцевый","глянцевая","прозрачный","прозрачная",
          "рыжий","рыжая","рыжие","оранжевый","оранжевая","оранжевые",
          "зелёный","зелёная","жёлтый","жёлтая","золотой","золотая","бежевый","бежевая",
          "коричневый","коричневая","фиолетовый","розовый","голубой"}
FILLER = {"и","или","а","но","же","ну","на","в","во","от","до","по","с","со","у","к","за","из","о","об",
          "нужно","нужен","нужна","нужны","хочу","заказ","заказать","заказываю","пожалуйста","добавьте",
          "добавь","примите","привезите","если","есть","нет","можно","надо","как","тоже","ещё","еще",
          "для","под","над","про",
          "там","тут","это","этот","эта","эти","то","тот","та","те","они","он","она","оно","мы","вы",
          "день","утро","вечер","завтра","сегодня","через","минут","минуты","часов","час",
          "добрый","доброе","добрая","здравствуйте","привет","спасибо","ок","хорошо","принято",
          "большой","большая","маленький","маленькая","новый","новая","много","мало","чуть",
          "пл","ст","мм","см","шт","м","кг","г","гр","год","лет","же","уже","или","тоже"}
STOP = UNITS_WORDS | COLORS | FILLER

def extract_key(raw: str):
    """Return (key, sizes_list, colors_list) from a raw_name."""
    raw = raw.lower().strip()
    sizes_compound = SIZE_RE.findall(raw)
    remaining = raw
    for s in sizes_compound:
        remaining = remaining.replace(s, " ")
    simple_nums = DIGITS_RE.findall(remaining)
    sizes = []
    seen = set()
    for s in sizes_compound + simple_nums:
        if s not in seen:
            seen.add(s); sizes.append(s)
    all_words = LETTERS_RE.findall(raw)
    colors_found = [w for w in all_words if w in COLORS]
    content = [w for w in all_words if w not in STOP and len(w) >= 2]
    if not content:
        content = [w for w in all_words if len(w) >= 2][:1]
    stems = sorted(set(stem(w) for w in content))[:4]
    return " ".join(stems), sizes, list(set(colors_found))

# ============= Load data =============

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()
cur.execute("""
    SELECT msg_id, raw_name, quantity, unit, confidence
    FROM extracted_order_items
    ORDER BY raw_name COLLATE NOCASE, msg_id
""")
all_rows = cur.fetchall()
print(f"Loaded {len(all_rows)} positions")

# ============= Dictionary (sheet 2) =============
agg = defaultdict(lambda: {"total": 0, "high": 0, "medium": 0, "low": 0, "units": Counter(), "msgs": set()})
for r in all_rows:
    key = (r["raw_name"] or "").strip().lower()
    if not key:
        continue
    a = agg[key]
    a["total"] += 1
    a[r["confidence"] or "low"] += 1
    u = (r["unit"] or "").strip() or "—"
    a["units"][u] += 1
    a["msgs"].add(r["msg_id"])

dict_rows = []
for name, a in sorted(agg.items(), key=lambda x: -x[1]["total"]):
    top_units = ", ".join(f"{u}:{c}" for u, c in a["units"].most_common(3))
    dict_rows.append({
        "raw_name": name, "total_count": a["total"], "unique_msgs": len(a["msgs"]),
        "high": a["high"], "medium": a["medium"], "low": a["low"], "top_units": top_units,
    })
print(f"Dictionary entries: {len(dict_rows)}")

# ============= Fuzzy groups (sheet 1) =============
name_cnt = Counter()
name_msgs = defaultdict(set)
for r in all_rows:
    n = (r["raw_name"] or "").strip().lower()
    if n:
        name_cnt[n] += 1
        name_msgs[n].add(r["msg_id"])

stem_groups = defaultdict(lambda: {"items": [], "sizes": Counter(), "colors": Counter(), "total": 0, "msgs": set(), "units": Counter()})
for name, cnt in name_cnt.items():
    key, sizes, colors = extract_key(name)
    if not key:
        key = "(empty)"
    g = stem_groups[key]
    g["items"].append((name, cnt))
    g["total"] += cnt
    g["msgs"].update(name_msgs[name])
    for s in sizes: g["sizes"][s] += cnt
    for c in colors: g["colors"][c] += cnt

# Unit counter per group
for r in all_rows:
    n = (r["raw_name"] or "").strip().lower()
    if not n: continue
    key, _, _ = extract_key(n)
    if not key: key = "(empty)"
    u = (r["unit"] or "").strip() or "—"
    stem_groups[key]["units"][u] += 1

# Direct-pair fuzzy merge (no chaining): small absorbs into nearest larger
all_keys_sorted = sorted(stem_groups.keys(), key=lambda k: -stem_groups[k]["total"])
single = [k for k in all_keys_sorted if " " not in k and len(k) >= 5]
merge_map = {}
for i, small in enumerate(single):
    small_total = stem_groups[small]["total"]
    for big in single[:i]:
        big_total = stem_groups[big]["total"]
        if big_total < small_total * 3:
            continue
        if abs(len(big) - len(small)) > 1:
            continue
        if lev(big, small) <= 1:
            merge_map[small] = big
            break

merged = defaultdict(lambda: {"items": [], "sizes": Counter(), "colors": Counter(), "total": 0, "msgs": set(), "units": Counter(), "stem_keys": set()})
for k, g in stem_groups.items():
    root = merge_map.get(k, k)
    m = merged[root]
    m["items"].extend(g["items"])
    m["total"] += g["total"]
    m["msgs"].update(g["msgs"])
    m["sizes"].update(g["sizes"])
    m["colors"].update(g["colors"])
    m["units"].update(g["units"])
    m["stem_keys"].add(k)

print(f"Groups: {len(merged)} (merged {len(merge_map)} typo-variants)")

group_rows = []
for key, g in sorted(merged.items(), key=lambda x: -x[1]["total"]):
    sorted_items = sorted(g["items"], key=lambda x: -x[1])
    top_members = " | ".join(f"{n}({c})" for n, c in sorted_items[:5])
    sizes_str = ", ".join(f"{s}:{c}" for s, c in g["sizes"].most_common(8))
    colors_str = ", ".join(f"{c}:{n}" for c, n in g["colors"].most_common(5))
    units_str = ", ".join(f"{u}:{c}" for u, c in g["units"].most_common(3))
    stem_keys_str = ", ".join(sorted(g["stem_keys"])) if len(g["stem_keys"]) > 1 else ""
    group_rows.append({
        "group_key": key,
        "total": g["total"],
        "variants": len(g["items"]),
        "unique_msgs": len(g["msgs"]),
        "top_members": top_members,
        "sizes": sizes_str,
        "colors": colors_str,
        "units": units_str,
        "merged_stems": stem_keys_str,
    })

# ============= Stats =============
total_items = len(all_rows)
total_msgs = len({r["msg_id"] for r in all_rows})
total_unique_names = len(dict_rows)
conf_dist = Counter(r["confidence"] for r in all_rows)
unit_dist = Counter((r["unit"] or "—") for r in all_rows)

# ============= Excel =============
wb = Workbook()

# Sheet 1: Группы (кластеры)
ws0 = wb.active
ws0.title = "Группы"
headers0 = ["Группа (стем)", "Всего", "Вариантов", "Уник msg", "Топ члены", "Размеры", "Цвета", "Единицы", "Слитые стемы"]
ws0.append(headers0)
for cell in ws0[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for r in group_rows:
    ws0.append([r["group_key"], r["total"], r["variants"], r["unique_msgs"],
                r["top_members"], r["sizes"], r["colors"], r["units"], r["merged_stems"]])
ws0.column_dimensions["A"].width = 28
ws0.column_dimensions["B"].width = 8
ws0.column_dimensions["C"].width = 10
ws0.column_dimensions["D"].width = 10
ws0.column_dimensions["E"].width = 70
ws0.column_dimensions["F"].width = 40
ws0.column_dimensions["G"].width = 25
ws0.column_dimensions["H"].width = 18
ws0.column_dimensions["I"].width = 25
ws0.freeze_panes = "A2"

# Sheet 2: Словарь
ws1 = wb.create_sheet("Словарь")
headers1 = ["raw_name", "Всего", "Уник msg", "high", "medium", "low", "Топ единицы"]
ws1.append(headers1)
for cell in ws1[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center")
for r in dict_rows:
    ws1.append([r["raw_name"], r["total_count"], r["unique_msgs"], r["high"], r["medium"], r["low"], r["top_units"]])
ws1.column_dimensions["A"].width = 50
ws1.column_dimensions["B"].width = 8
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 8
ws1.column_dimensions["E"].width = 10
ws1.column_dimensions["F"].width = 8
ws1.column_dimensions["G"].width = 30
ws1.freeze_panes = "A2"

# Sheet 3: Все позиции
ws2 = wb.create_sheet("Все позиции")
headers2 = ["msg_id", "raw_name", "quantity", "unit", "confidence"]
ws2.append(headers2)
for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center")
for r in all_rows:
    ws2.append([r["msg_id"], r["raw_name"], r["quantity"], r["unit"], r["confidence"]])
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 50
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12
ws2.freeze_panes = "A2"

# Sheet 4: Статистика
ws3 = wb.create_sheet("Статистика")
ws3.append(["Метрика", "Значение"])
for cell in ws3[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5496")

stats = [
    ("Всего извлечённых позиций", total_items),
    ("Уникальных сообщений с заказами", total_msgs),
    ("Уникальных raw_name (после lower+trim)", total_unique_names),
    ("Кластеров после фуззи-группировки", len(group_rows)),
    ("", ""),
    ("Confidence high", conf_dist.get("high", 0)),
    ("Confidence medium", conf_dist.get("medium", 0)),
    ("Confidence low", conf_dist.get("low", 0)),
    ("", ""),
    ("Кандидатов в order_candidates", 15702),
    ("Доля msg с заказами от кандидатов", f"{100 * total_msgs / 15702:.1f}%"),
    ("", ""),
]
for label, val in stats:
    ws3.append([label, val])
ws3.append(["", ""])
ws3.append(["Топ-10 единиц измерения", ""])
for u, c in unit_dist.most_common(10):
    ws3.append([u, c])
ws3.column_dimensions["A"].width = 45
ws3.column_dimensions["B"].width = 18

xlsx_path = f"{OUT_DIR}/order_items_dictionary.xlsx"
wb.save(xlsx_path)
print(f"Saved {xlsx_path}")

# ============= CSV =============
csv_groups = f"{OUT_DIR}/order_items_groups.csv"
with open(csv_groups, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["group_key", "total", "variants", "unique_msgs", "top_members", "sizes", "colors", "units", "merged_stems"])
    for r in group_rows:
        w.writerow([r["group_key"], r["total"], r["variants"], r["unique_msgs"],
                    r["top_members"], r["sizes"], r["colors"], r["units"], r["merged_stems"]])
print(f"Saved {csv_groups}")

csv_dict = f"{OUT_DIR}/order_items_dictionary.csv"
with open(csv_dict, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["raw_name", "total_count", "unique_msgs", "high", "medium", "low", "top_units"])
    for r in dict_rows:
        w.writerow([r["raw_name"], r["total_count"], r["unique_msgs"], r["high"], r["medium"], r["low"], r["top_units"]])
print(f"Saved {csv_dict}")

csv_raw = f"{OUT_DIR}/order_items_raw.csv"
with open(csv_raw, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["msg_id", "raw_name", "quantity", "unit", "confidence"])
    for r in all_rows:
        w.writerow([r["msg_id"], r["raw_name"], r["quantity"], r["unit"], r["confidence"]])
print(f"Saved {csv_raw}")

print(f"\nFINAL: {total_items} positions, {total_msgs} msgs, {total_unique_names} unique names, {len(group_rows)} groups")
conn.close()
